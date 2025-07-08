import os
import time
import subprocess
import shutil
from typing import List, Dict, Any
from pdf2image import convert_from_path
from PIL import Image
import base64
from io import BytesIO
from config import Config
import fitz  # PyMuPDF
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import win32com.client as win32
import pythoncom

class PDFProcessor:
    def __init__(self):
        self.temp_dir = Config.TEMP_DIR
        os.makedirs(self.temp_dir, exist_ok=True)

    def to_cmd_path(self, path):
        return os.path.normpath(path).replace('\\', '/')

    def docx_to_pdf(self, docx_path: str) -> str:
        """Convert docx file to PDF using LibreOffice"""
        base_name = os.path.splitext(os.path.basename(docx_path))[0]
        pdf_path = os.path.join(self.temp_dir, f"{base_name}_{int(time.time())}.pdf")
        pdf_path = self.to_cmd_path(pdf_path)
        docx_path_cmd = self.to_cmd_path(docx_path)
        temp_dir_cmd = self.to_cmd_path(self.temp_dir)
        soffice_path = self.to_cmd_path(os.path.join('C:', 'Program Files', 'LibreOffice', 'program', 'soffice.exe'))
        
        print(f"Converting {docx_path} to PDF using LibreOffice...")
        
        try:
            subprocess.run([
                soffice_path,
                '--headless',
                '--convert-to', 'pdf',
                docx_path_cmd,
                '--outdir', temp_dir_cmd
            ], check=True)
            
            generated_pdf = os.path.join(self.temp_dir, f"{base_name}.pdf")
            generated_pdf = self.to_cmd_path(generated_pdf)
            if os.path.exists(generated_pdf):
                shutil.move(generated_pdf, pdf_path)
                print(f"PDF conversion completed: {pdf_path}")
                return pdf_path
            else:
                raise Exception("LibreOffice did not generate the expected PDF file")
                
        except Exception as e:
            print(f"PDF conversion failed: {e}")
            raise

    def excel_to_pdf(self, excel_path: str) -> str:
        """Convert Excel file to PDF using COM automation"""
        try:
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            pdf_path = os.path.join(self.temp_dir, f"{base_name}_{int(time.time())}.pdf")
            
            print(f"Converting {excel_path} to PDF...")
            
            # Initialize COM
            pythoncom.CoInitialize()
            
            try:
                # Create Excel application
                excel_app = win32.Dispatch("Excel.Application")
                excel_app.Visible = False
                excel_app.DisplayAlerts = False
                
                # Open workbook
                workbook = excel_app.Workbooks.Open(os.path.abspath(excel_path))
                
                # Export as PDF
                workbook.ExportAsFixedFormat(0, pdf_path)
                
                # Close workbook and quit Excel
                workbook.Close()
                excel_app.Quit()
                
            finally:
                pythoncom.CoUninitialize()
            
            print(f"PDF conversion completed: {pdf_path}")
            return pdf_path
            
        except Exception as e:
            print(f"Excel to PDF conversion failed: {e}")
            # Fallback: create a simple text-based representation
            return self._excel_to_pdf_fallback(excel_path)

    def _excel_to_pdf_fallback(self, excel_path: str) -> str:
        """Fallback method for Excel to PDF conversion"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import inch
            
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            pdf_path = os.path.join(self.temp_dir, f"{base_name}_{int(time.time())}_fallback.pdf")
            
            # Create PDF with Excel content as text
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter
            
            # Load Excel content
            wb = openpyxl.load_workbook(excel_path)
            y_position = height - inch
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                c.drawString(inch, y_position, f"Sheet: {sheet_name}")
                y_position -= 20
                
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        c.drawString(inch, y_position, row_text[:80])  # Limit line length
                        y_position -= 15
                        
                        if y_position < inch:  # New page if needed
                            c.showPage()
                            y_position = height - inch
            
            c.save()
            wb.close()
            
            print(f"Fallback PDF conversion completed: {pdf_path}")
            return pdf_path
            
        except Exception as e:
            print(f"Fallback PDF conversion also failed: {e}")
            raise

    def document_to_pdf(self, file_path: str) -> str:
        """Convert document to PDF - supports both Word and Excel"""
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.docx':
            return self.docx_to_pdf(file_path)
        elif file_ext in ['.xlsx', '.xls']:
            return self.excel_to_pdf(file_path)
        else:
            raise ValueError(f"Unsupported file format for PDF conversion: {file_ext}")

    def pdf_to_images(self, pdf_path: str, dpi: int = 200) -> List[Dict[str, Any]]:
        """Use PyMuPDF to convert each page of PDF to an image"""
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        page_images = []
        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            page = doc[page_num]
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            img_filename = os.path.join(self.temp_dir, f"{base_name}_{page_num+1}.png")
            pix.save(img_filename)
            page_images.append({
                'page_number': page_num + 1,
                'image_path': img_filename,
                'image_format': 'PNG',
                'dpi': dpi
            })
        doc.close()
        print(f"PDF to image conversion completed, total {len(page_images)} pages")
        return page_images

    def process_document_with_images(self, file_path: str) -> Dict[str, Any]:
        """Process document: convert to PDF and generate page screenshots"""
        try:
            pdf_path = self.document_to_pdf(file_path)
            page_images = self.pdf_to_images(pdf_path)
            
            return {
                'pdf_path': pdf_path,
                'page_images': page_images,
                'total_pages': len(page_images)
            }
            
        except Exception as e:
            print(f"Document processing failed: {e}")
            raise

    def cleanup_temp_files(self, pdf_path: str = None):
        """Clean up temporary files"""
        try:
            if pdf_path and os.path.exists(pdf_path):
                os.remove(pdf_path)
                print(f"Temporary PDF file deleted: {pdf_path}")
        except Exception as e:
            print(f"Failed to clean up temporary files: {e}")

    def extract_text_from_pdf(self, pdf_path: str) -> str:
        """
        Extract text content from PDF file
        Args:
            pdf_path: PDF file path
        Returns:
            Extracted text content
        """
        if not os.path.exists(pdf_path):
            print(f"PDF file not found: {pdf_path}")
            return ""
        
        try:
            doc = fitz.open(pdf_path)
            text_content = ""
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                page_text = page.get_text()
                if page_text.strip():
                    text_content += f"\n--- Page {page_num + 1} ---\n"
                    text_content += page_text.strip()
                    text_content += "\n"
            
            doc.close()
            
            if text_content.strip():
                print(f"Successfully extracted text from PDF, total {len(doc)} pages")
                return text_content.strip()
            else:
                print("No extractable text content found in PDF file")
                return ""
                
        except Exception as e:
            print(f"Failed to extract text from PDF: {e}")
            return "" 